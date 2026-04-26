"""Baker Hughes North America Rotary Rig Count ingester.

Source: Baker Hughes publishes a "North America Rotary Rig Count - Pivot
Table" XLSX archive. The canonical landing page is
``https://rigcount.bakerhughes.com/na-rig-count`` and the actual XLSX
URL has the form ``https://rigcount.bakerhughes.com/static-files/<id>``,
where ``<id>`` is an unstable Sitecore content id. Because the URL
rotates, this ingester accepts EITHER:

    * ``archive_url`` — explicit URL to the XLSX (operator overrides),
    * ``manual_xlsx_path`` — a manually-downloaded XLSX on disk.

If neither is provided, :class:`BakerHughesURLNotConfiguredError` is
raised, pointing the operator at the runbook stub
``docs/v2/operator_runbook_public_data.md``. Both paths are NO-API-KEY
public retrievals; rights status is ``public`` and the registry has
``model_eligible: true`` for the five extracted series.

PIT discipline:
    * ``release_ts`` for each weekly observation = Friday 13:00 ET in
      UTC. Baker Hughes releases the rig count Fridays at ~13:00 ET.
    * One :class:`v2.ingest.base.FetchResult` per series. The series
      DataFrame carries ``[observation_date, value, units, retrieved_at_utc]``.
      ``release_ts`` on the FetchResult is the *latest* publication
      timestamp in the series (most-recent Friday 13:00 ET in UTC).
    * ``revision_ts`` is left None — historical rows are immutable
      under the BH publication contract; if a row's value ever changes
      the writer's checksum-mismatch path will treat it as a revision.

The five emitted series are:

    * ``us_oil_total`` — sum of US Oil rigs (Land + Offshore).
    * ``us_gas_total`` — US Gas rigs.
    * ``us_total`` — US Total rigs.
    * ``canada_total`` — Canada Total rigs.
    * ``na_total`` — North America Total rigs.
"""

from __future__ import annotations

import io
import re
import tempfile
from datetime import UTC, datetime, time
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook

from v2.ingest._http import HTTPClient
from v2.ingest.base import BaseIngester, FetchResult
from v2.pit_store.manifest import PITManifest
from v2.pit_store.writer import PITWriter

SCRAPER_VERSION = "v2.b2b.0"
NY_TZ = ZoneInfo("America/New_York")
BH_RELEASE_LOCAL_TIME = time(13, 0)  # 13:00 ET
OPERATOR_RUNBOOK_REF = "docs/v2/operator_runbook_public_data.md"

SERIES_ORDER: tuple[str, ...] = (
    "us_oil_total",
    "us_gas_total",
    "us_total",
    "canada_total",
    "na_total",
)


class BakerHughesURLNotConfiguredError(RuntimeError):
    """Neither ``archive_url`` nor ``manual_xlsx_path`` was provided.

    Carries the operator runbook reference so the on-call can either
    set the rotating XLSX URL or supply a manually-downloaded file.
    """

    def __init__(self, operator_runbook_ref: str = OPERATOR_RUNBOOK_REF) -> None:
        super().__init__(
            "BakerHughesIngester requires either archive_url or "
            f"manual_xlsx_path. See operator runbook: {operator_runbook_ref}"
        )
        self.operator_runbook_ref = operator_runbook_ref


class BakerHughesXLSXParseError(ValueError):
    """The XLSX could not be parsed into the expected pivot-table shape."""


def _release_ts_for(d: pd.Timestamp | datetime) -> datetime:
    """Return Friday 13:00 ET (in UTC) for the observation's week.

    Baker Hughes publishes weekly on Friday afternoon. We pin
    ``release_ts`` to the row's date at 13:00 America/New_York and
    convert to UTC.
    """
    py_dt = d.to_pydatetime() if isinstance(d, pd.Timestamp) else d
    local = datetime.combine(py_dt.date(), BH_RELEASE_LOCAL_TIME, tzinfo=NY_TZ)
    return local.astimezone(UTC)


class BakerHughesIngester(BaseIngester):
    """Baker Hughes North America rig count ingester."""

    name = "baker_hughes_rig_count"
    source = "baker_hughes_rig_count"

    def __init__(
        self,
        writer: PITWriter,
        manifest: PITManifest,
        *,
        http: HTTPClient | None = None,
        archive_url: str | None = None,
        manual_xlsx_path: Path | None = None,
    ) -> None:
        super().__init__(writer, manifest)
        if archive_url is None and manual_xlsx_path is None:
            raise BakerHughesURLNotConfiguredError(OPERATOR_RUNBOOK_REF)
        self._archive_url = archive_url
        self._manual_xlsx_path = (
            Path(manual_xlsx_path) if manual_xlsx_path is not None else None
        )
        self._http_owned = http is None and archive_url is not None
        self._http = http if http is not None else (
            HTTPClient() if archive_url is not None else None
        )

    # -- public --------------------------------------------------------------

    def fetch(self, as_of_ts: datetime | None = None) -> list[FetchResult]:
        retrieved_at, method, xlsx_bytes = self._load_xlsx_bytes()
        df = self._parse_xlsx(xlsx_bytes)

        if df.empty:
            raise BakerHughesXLSXParseError(
                "Baker Hughes XLSX produced an empty dataframe"
            )

        latest_obs_date = df["observation_date"].max()
        latest_release_ts = _release_ts_for(latest_obs_date)
        observation_start = df["observation_date"].min()
        observation_end = latest_obs_date

        results: list[FetchResult] = []
        for series in SERIES_ORDER:
            sub = pd.DataFrame(
                {
                    "observation_date": df["observation_date"].dt.date,
                    "value": df[series].astype("float64"),
                    "units": ["count"] * len(df),
                    "retrieved_at_utc": pd.Series(
                        [retrieved_at] * len(df),
                        dtype="datetime64[ns, UTC]",
                    ),
                }
            )
            provenance: dict[str, Any] = {
                "source": self.source,
                "method": method,
                "scraper_version": SCRAPER_VERSION,
                "series": series,
                "endpoint": self._archive_url,
                "manual_xlsx_path": (
                    str(self._manual_xlsx_path)
                    if self._manual_xlsx_path is not None
                    else None
                ),
                "retrieved_at_utc": retrieved_at.isoformat(),
            }
            results.append(
                FetchResult(
                    source=self.source,
                    series=series,
                    release_ts=latest_release_ts,
                    revision_ts=None,
                    data=sub,
                    provenance=provenance,
                    observation_start=(
                        observation_start.date()
                        if hasattr(observation_start, "date")
                        else observation_start
                    ),
                    observation_end=(
                        observation_end.date()
                        if hasattr(observation_end, "date")
                        else observation_end
                    ),
                )
            )
        return results

    def close(self) -> None:
        if self._http_owned and self._http is not None:
            self._http.close()

    # -- internals -----------------------------------------------------------

    def _load_xlsx_bytes(self) -> tuple[datetime, str, bytes]:
        if self._manual_xlsx_path is not None:
            data = self._manual_xlsx_path.read_bytes()
            return datetime.now(UTC), "manual_xlsx", data
        assert self._archive_url is not None
        assert self._http is not None
        resp = self._http.get(self._archive_url)
        if resp.status_code != 200:
            raise BakerHughesXLSXParseError(
                f"Baker Hughes archive HTTP {resp.status_code} for "
                f"{self._archive_url!r}"
            )
        return resp.retrieved_at_utc, "xlsx_download", resp.content

    @staticmethod
    def _parse_xlsx(payload: bytes) -> pd.DataFrame:
        """Parse the BH "Pivot Table" XLSX into a wide-form dataframe.

        Returns columns:
            observation_date, us_oil_total, us_gas_total,
            us_total, canada_total, na_total
        """
        # Use openpyxl in read-only / values-only mode for resilience to
        # the various sheet shapes BH has used over time. Operate on
        # bytes via a tmp file (openpyxl prefers a path).
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            tmp.write(payload)
            tmp.flush()
            wb = load_workbook(tmp.name, read_only=True, data_only=True)
            try:
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            finally:
                wb.close()

        if not rows:
            return pd.DataFrame(
                columns=list(("observation_date",) + SERIES_ORDER)
            )

        header = [_norm_header(c) for c in rows[0]]
        body = rows[1:]

        col_index = _resolve_columns(header)

        records: list[dict[str, Any]] = []
        for r in body:
            if r is None or all(c is None for c in r):
                continue
            try:
                date_cell = r[col_index["date"]]
            except IndexError:
                continue
            if date_cell is None:
                continue
            try:
                obs_date = pd.Timestamp(date_cell)
            except (ValueError, TypeError):
                continue
            if pd.isna(obs_date):
                continue

            us_oil_land = _to_float(r, col_index.get("us_oil_land"))
            us_oil_offshore = _to_float(r, col_index.get("us_oil_offshore"))
            us_oil_total_direct = _to_float(r, col_index.get("us_oil_total"))
            if us_oil_total_direct is not None:
                us_oil_total = us_oil_total_direct
            else:
                us_oil_total = (us_oil_land or 0.0) + (us_oil_offshore or 0.0)

            rec = {
                "observation_date": obs_date,
                "us_oil_total": us_oil_total,
                "us_gas_total": _to_float(r, col_index.get("us_gas_total")),
                "us_total": _to_float(r, col_index.get("us_total")),
                "canada_total": _to_float(r, col_index.get("canada_total")),
                "na_total": _to_float(r, col_index.get("na_total")),
            }
            records.append(rec)

        df = pd.DataFrame.from_records(records)
        if df.empty:
            return df
        df = df.sort_values("observation_date").reset_index(drop=True)
        return df


# -- sheet-shape helpers -----------------------------------------------------


_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("date", "publishdate", "weekend", "weekendingdate"),
    "us_oil_land": ("usoilland", "usoillandrigs", "uslandoil"),
    "us_oil_offshore": (
        "usoiloffshore",
        "usoiloffshorerigs",
        "usoffshoreoil",
    ),
    "us_oil_total": ("usoiltotal", "usoil", "usoilrigs"),
    "us_gas_total": ("usgastotal", "usgas", "usgasrigs"),
    "us_total": ("ustotal", "ustotalrigs"),
    "canada_total": ("canadatotal", "canada", "canadarigs", "canadatotalrigs"),
    "na_total": ("natotal", "northamericatotal", "natotalrigs"),
}


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    # Strip non-alphanumerics for tolerant matching.
    return re.sub(r"[^a-z0-9]+", "", s)


def _resolve_columns(header: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for canonical, aliases in _HEADER_ALIASES.items():
        for i, h in enumerate(header):
            if h in aliases:
                out[canonical] = i
                break
    if "date" not in out:
        raise BakerHughesXLSXParseError(
            f"Baker Hughes XLSX: cannot locate date column in header={header!r}"
        )
    required = ("us_gas_total", "us_total", "canada_total", "na_total")
    missing = [r for r in required if r not in out]
    if missing:
        raise BakerHughesXLSXParseError(
            f"Baker Hughes XLSX: missing required columns {missing!r}; "
            f"header={header!r}"
        )
    if "us_oil_total" not in out and not (
        "us_oil_land" in out or "us_oil_offshore" in out
    ):
        raise BakerHughesXLSXParseError(
            "Baker Hughes XLSX: need either us_oil_total OR "
            f"(us_oil_land and us_oil_offshore); header={header!r}"
        )
    return out


def _to_float(row: tuple[Any, ...], idx: int | None) -> float | None:
    if idx is None:
        return None
    try:
        v = row[idx]
    except IndexError:
        return None
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# Ensure unused-import lint stays clean (io retained for symmetry with other
# ingesters that may stream into pandas).
_ = io


__all__ = [
    "OPERATOR_RUNBOOK_REF",
    "SERIES_ORDER",
    "BakerHughesIngester",
    "BakerHughesURLNotConfiguredError",
    "BakerHughesXLSXParseError",
]
